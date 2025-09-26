"""Flask app factory for the dryer-dashboard project.

This package exposes create_app() so you can run the app with the Flask CLI:

  FLASK_APP=app:create_app flask run

The app serves a small API at /api/sensors (returns JSON) and serves
the static `index.html` at the root.
"""
from flask import Flask, jsonify
import os
from pathlib import Path
from datetime import datetime

from . import sensors as sensors_module
from .logger import create_logger

# instantiate a CSV logger (logs directory inside project static area)
LOG_DIR = str(project_root := Path(__file__).resolve().parent.parent / 'logs') if 'project_root' not in globals() else str(project_root / 'logs')
logger = create_logger(LOG_DIR, interval_seconds=int(os.environ.get('LOG_INTERVAL', 900)))

# Simple in-memory auger state. This is deliberately minimal: it keeps the
# current discharge auger percentage (0..100). For persistence across reboots
# or multiple processes you should move this to a small database or expose it
# via a hardware controller. Kept here for quick UI control during testing.
_AUGER_PCT = 50


def create_app():
    # Serve static files from the top-level `static/` directory so the
    # kiosk UI (located at ../static/index.html) is available at '/'. Using
    # an absolute path avoids issues when the working directory changes.
    project_root = Path(__file__).resolve().parent.parent
    static_dir = str(project_root / 'static')
    app = Flask(__name__, static_folder=static_dir, static_url_path="/")

    # Create logger instance after project_root is known
    log_dir = str(project_root / 'logs')
    from .logger import create_logger
    _logger = create_logger(log_dir, interval_seconds=int(os.environ.get('LOG_INTERVAL', 900)))

    @app.route("/api/sensors")
    def api_sensors():
        """Return the latest sensor readings as JSON.

        Prefer a sensors.read_all() helper when the module provides it. If not
        available we call get_temps()/get_moisture() and build a small JSON
        payload so the endpoint remains usable across different sensor modules.
        """
        if hasattr(sensors_module, "read_all"):
            try:
                return jsonify(sensors_module.read_all())
            except Exception:
                # Fall through to a safer fallback below and return whatever we can
                pass

        # Fallback: build a minimal response from available getters
        try:
            inlet_c, outlet_c = sensors_module.get_temps(return_fahrenheit=False)
        except Exception:
            inlet_c = outlet_c = None

        try:
            inlet_v, outlet_v = sensors_module.get_moisture()
        except Exception:
            inlet_v = outlet_v = None

        data = {
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "inlet_c": inlet_c,
            "outlet_c": outlet_c,
            "inlet_v": inlet_v,
            "outlet_v": outlet_v,
        }

        return jsonify(data)


    @app.route("/data")
    def data_for_ui():
        """Return transformed sensor data tailored to the kiosk UI.

        Fields returned (example):
          - temp_in_c, temp_out_c (floats or null)
          - temp_in_f, temp_out_f (floats or null)
          - moisture_in (0-100 %), moisture_out (0-100 %)
          - bushels_per_hr (float) -- crude estimate for now
          - timestamp
        """
        try:
            payload = sensors_module.read_all() if hasattr(sensors_module, 'read_all') else {}
        except Exception:
            payload = {}

        inlet_c = payload.get('inlet_c')
        outlet_c = payload.get('outlet_c')
        inlet_v = payload.get('inlet_v')
        outlet_v = payload.get('outlet_v')

        def c_to_f(c):
            if c is None:
                return None
            return (c * 9 / 5) + 32

        # Convert voltage (0..3.3) to percent (0..100). Clamp defensively.
        def v_to_pct(v):
            try:
                pct = (float(v) / 3.3) * 100.0
                if pct < 0:
                    pct = 0.0
                if pct > 100:
                    pct = 100.0
                return pct
            except Exception:
                return None

        moisture_in = v_to_pct(inlet_v)
        moisture_out = v_to_pct(outlet_v)

        # Very rough bushels/hr estimate: linear scale from dry -> full rate.
        # This is a placeholder and should be replaced with real calibration.
        if moisture_in is None:
            bushels = None
        else:
            bushels = max(0.0, (1.0 - (moisture_in / 100.0)) * 100.0)

        # Determine per-sensor health flags from the sensors payload. If the
        # sensors module provides an `errors` list we use that to mark inlet
        # and outlet as healthy/unhealthy. Otherwise we default to True when
        # readings look present.
        errs = payload.get('errors') if isinstance(payload, dict) else None
        simulated = bool(payload.get('simulated')) if isinstance(payload, dict) else False

        if isinstance(errs, list) and errs:
            inlet_ok = not any('inlet' in str(e).lower() for e in errs)
            outlet_ok = not any('outlet' in str(e).lower() for e in errs)
        else:
            # If no explicit errors, consider a sensor OK when we have a value
            inlet_ok = inlet_c is not None or inlet_v is not None
            outlet_ok = outlet_c is not None or outlet_v is not None

        data = {
            'timestamp': payload.get('timestamp'),
            'temp_in_c': inlet_c,
            'temp_out_c': outlet_c,
            'temp_in_f': None if inlet_c is None else c_to_f(inlet_c),
            'temp_out_f': None if outlet_c is None else c_to_f(outlet_c),
            'moisture_in': moisture_in,
            'moisture_out': moisture_out,
            'bushels_per_hr': bushels,
            'auger_pct': globals().get('_AUGER_PCT', None),
            'inlet_ok': bool(inlet_ok),
            'outlet_ok': bool(outlet_ok),
            'simulated': simulated,
        }
        return jsonify(data)


    @app.route('/status')
    def ui_status():
        """Return a minimal status object for the UI (e.g. recording state).

        We mark `recording` True when sensors returned no errors in read_all().
        """
        recording = True
        try:
            payload = sensors_module.read_all() if hasattr(sensors_module, 'read_all') else {}
            errs = payload.get('errors') if isinstance(payload, dict) else None
            if errs:
                recording = False
        except Exception:
            recording = False

        return jsonify({'recording': recording})

    @app.route('/auger', methods=['GET', 'POST'])
    def auger_control():
        """Simple endpoint to get/set the discharge auger percentage.

        GET  -> {'auger_pct': number}
        POST -> accepts JSON {delta: number} to change current value by delta
                (positive or negative). Returns {'auger_pct': new_value}.
        """
        nonlocal_vars = globals()
        try:
            cur = int(nonlocal_vars.get('_AUGER_PCT', 50))
        except Exception:
            cur = 50

        from flask import request
        if request.method == 'POST':
            try:
                j = request.get_json(force=True)
                delta = int(j.get('delta', 0))
            except Exception:
                delta = 0

            new = max(0, min(100, cur + delta))
            nonlocal_vars['_AUGER_PCT'] = new
            return jsonify({'auger_pct': new})

        # GET
        return jsonify({'auger_pct': cur})

    # Logging control endpoints
    @app.route('/logs/start', methods=['POST'])
    def logs_start():
        try:
            _logger.start()
            return jsonify({'running': True, 'file': _logger.current_file()})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/logs/stop', methods=['POST'])
    def logs_stop():
        try:
            _logger.stop()
            return jsonify({'running': False})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/logs/status')
    def logs_status():
        return jsonify({'running': _logger.is_running(), 'file': _logger.current_file()})

    @app.route('/logs/list')
    def logs_list():
        return jsonify({'logs': _logger.list_logs()})

    @app.route('/logs/sample', methods=['POST'])
    def logs_sample():
        """Trigger a one-off sample and append to the current log file. If no log exists,
        a new file will be created. Returns the file name written to or an error."""
        try:
            name = _logger.sample_once()
            if name is None:
                return jsonify({'error': 'failed to write sample'}), 500
            return jsonify({'file': name})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/logs/latest')
    def logs_latest():
        """Return the most recent data row from the newest log file as plain text."""
        try:
            row = _logger.get_latest_row()
            if row is None:
                return jsonify({'row': None})
            return jsonify({'row': row})
        except Exception as e:
            return jsonify({'error': str(e)}), 500


    @app.route('/logs/preview')
    def logs_preview():
        """Return the last N parsed rows from the newest log file as JSON.

        Query params:
          - limit (int) default 10
        """
        from flask import request
        from zoneinfo import ZoneInfo
        UTC = ZoneInfo('UTC')
        LOCAL = ZoneInfo('America/Chicago')
        try:
            limit = int(request.args.get('limit', 10))
        except Exception:
            limit = 10

        files = _logger.list_logs()
        if not files:
            return jsonify({'rows': []})

        newest = Path(log_dir) / files[0]

        import csv, ast
        from datetime import datetime

        try:
            with newest.open('r', newline='') as f:
                reader = csv.reader(f)
                all_rows = list(reader)
        except Exception:
            return jsonify({'rows': []})

        if not all_rows or len(all_rows) < 2:
            return jsonify({'rows': []})

        header = all_rows[0]
        header_map = {
            'timestamp': 'Timestamp',
            'inlet_c': 'Temp In',
            'outlet_c': 'Temp Out',
            'inlet_v': 'Moisture In',
            'outlet_v': 'Moisture Out',
            'simulated': 'Simulated',
            'errors': 'Errors',
            'auger_pct': 'Discharge Percentage',
            'bushels_per_hr': 'Bushels-per-hr',
        }
        friendly_keys = [header_map.get(h, h) for h in header]

        def try_parse(value):
            if value is None or value == '':
                return None
            try:
                if '.' in value:
                    return float(value)
                return int(value)
            except Exception:
                pass
            try:
                s = value
                # parse as UTC and convert to local time
                if isinstance(s, str) and s.endswith('Z'):
                    s2 = s[:-1]
                    dt = datetime.fromisoformat(s2)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=UTC)
                    local_dt = dt.astimezone(LOCAL)
                    return local_dt.isoformat()
                dt = datetime.fromisoformat(s)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=UTC)
                local_dt = dt.astimezone(LOCAL)
                return local_dt.isoformat()
            except Exception:
                return value

        rows = []
        data_rows = all_rows[1:]
        for row in data_rows[-limit:]:
            obj = {}
            for i, val in enumerate(row):
                key = friendly_keys[i] if i < len(friendly_keys) else f'col{i}'
                if i < len(header) and header[i] == 'errors':
                    pretty = None
                    try:
                        parsed_errors = ast.literal_eval(val) if val not in (None, '') else []
                        if isinstance(parsed_errors, (list, tuple)):
                            pretty = '; '.join(str(x) for x in parsed_errors) if parsed_errors else None
                        else:
                            pretty = str(parsed_errors) if parsed_errors is not None else None
                    except Exception:
                        pretty = val
                    obj[key] = pretty
                else:
                    parsed = try_parse(val)
                    # convert datetime to ISO string for JSON
                    if hasattr(parsed, 'isoformat'):
                        obj[key] = parsed.isoformat()
                    else:
                        obj[key] = parsed
            rows.append(obj)

        return jsonify({'rows': rows})

    @app.route('/kiosk/exit', methods=['POST'])
    def kiosk_exit():
        """Attempt to stop the systemd kiosk service. Only callable from localhost."""
        from flask import request
        # simple origin check: only allow local calls
        if request.remote_addr not in ('127.0.0.1', '::1'):
            return jsonify({'error': 'forbidden'}), 403

        import subprocess
        try:
            subprocess.check_call(['systemctl', '--user', 'stop', 'dryer-kiosk.service'])
            return jsonify({'stopped': True})
        except subprocess.CalledProcessError:
            # try system service
            try:
                subprocess.check_call(['sudo', 'systemctl', 'stop', 'dryer-kiosk.service'])
                return jsonify({'stopped': True})
            except Exception as e:
                return jsonify({'error': str(e)}), 500

    @app.route('/logs/download/<path:name>')
    def logs_download(name):
        from flask import send_from_directory
        try:
            return send_from_directory(log_dir, name, as_attachment=True)
        except Exception:
            return jsonify({'error': 'not found'}), 404


    @app.route('/logs/export')
    def logs_export():
        """Export the newest CSV log as an XLSX file and return as attachment.

        Requires `openpyxl` to be installed. If not present, returns 503 with
        instructions for installing the dependency.
        """
        # find newest CSV
        files = _logger.list_logs()
        if not files:
            return jsonify({'error': 'no logs available'}), 404

        newest = Path(log_dir) / files[0]

        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception:
            return jsonify({'error': 'openpyxl not installed; run `pip install openpyxl`'}), 503

        # read CSV rows and produce typed XLSX for readability
        wb = Workbook()
        ws = wb.active

        import csv
        from datetime import datetime

        rows = []
        with newest.open('r', newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)

        if not rows:
            return jsonify({'error': 'empty log file'}), 404

        # header styling
        header = rows[0]
        # map internal CSV header names to friendly column names requested by user
        header_map = {
            'timestamp': 'Timestamp',
            'inlet_c': 'Temp In',
            'outlet_c': 'Temp Out',
            'inlet_v': 'Moisture In',
            'outlet_v': 'Moisture Out',
            'simulated': 'Simulated',
            'errors': 'Errors',
            'auger_pct': 'Discharge Percentage',
            'bushels_per_hr': 'Bushels-per-hr',
        }
        friendly_header = [header_map.get(h, h) for h in header]
        for ci, col in enumerate(friendly_header, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = Font(bold=True)

        # helper to try parse numeric or datetime
        from zoneinfo import ZoneInfo
        UTC = ZoneInfo('UTC')
        LOCAL = ZoneInfo('America/Chicago')

        def try_parse(value):
            if value is None or value == '':
                return None
            # try int/float
            try:
                if '.' in value:
                    return float(value)
                return int(value)
            except Exception:
                pass
            # try ISO datetime (strip trailing Z)
            try:
                s = value
                # if the string ends with Z, treat as UTC
                if isinstance(s, str) and s.endswith('Z'):
                    s2 = s[:-1]
                    dt = datetime.fromisoformat(s2)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=UTC)
                    # convert to local tz and return naive local datetime for Excel
                    local_dt = dt.astimezone(LOCAL)
                    return local_dt.replace(tzinfo=None)

                dt = datetime.fromisoformat(s)
                # if no tzinfo, assume UTC then convert
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=UTC)
                local_dt = dt.astimezone(LOCAL)
                return local_dt.replace(tzinfo=None)
            except Exception:
                return value

        # write data rows typed, with pretty-printing for errors column
        col_widths = [len(c) for c in header]
        # find index of 'errors' column in original CSV header
        try:
            errors_idx = header.index('errors')
        except ValueError:
            errors_idx = None

        import ast
        for ri, row in enumerate(rows[1:], start=2):
            for ci, val in enumerate(row, start=1):
                # pretty-print errors column if present
                if errors_idx is not None and (ci - 1) == errors_idx:
                    pretty = None
                    try:
                        parsed_errors = ast.literal_eval(val) if val not in (None, '') else []
                        if isinstance(parsed_errors, (list, tuple)):
                            pretty = '; '.join(str(x) for x in parsed_errors) if parsed_errors else None
                        else:
                            pretty = str(parsed_errors) if parsed_errors is not None else None
                    except Exception:
                        pretty = val
                    ws.cell(row=ri, column=ci, value=pretty)
                    text = str(pretty) if pretty is not None else ''
                else:
                    parsed = try_parse(val)
                    cell = ws.cell(row=ri, column=ci, value=parsed)
                    # if this is the timestamp column, apply a datetime format
                    try:
                        if header[ci-1] == 'timestamp' and isinstance(parsed, datetime):
                            cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    except Exception:
                        pass
                    text = str(parsed) if parsed is not None else ''

                # track width
                if ci-1 >= len(col_widths):
                    col_widths.append(len(text))
                else:
                    col_widths[ci-1] = max(col_widths[ci-1], len(text))

        # auto-size columns (simple heuristic)
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(w, 10), 50)

        out_name = newest.stem + '.xlsx'
        out_path = Path(log_dir) / out_name
        wb.save(out_path)

        from flask import send_from_directory
        return send_from_directory(log_dir, out_name, as_attachment=True)

    @app.route("/")
    def index():
        return app.send_static_file("index.html")

    return app
